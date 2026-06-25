"""Build specification spreadsheets (XLSX / CSV) for sales handoff."""

from __future__ import annotations

import csv
import io
import zipfile
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.core.datetime_utils import format_utc_as_moscow
from app.models.configuration import Configuration
from app.models.configuration_item import ConfigurationItem
from app.models.license import License
from app.models.module import Module
from app.models.product import Product

ExportLocale = Literal["ru", "en"]

_LOCALE_TEXT: dict[ExportLocale, dict[str, Any]] = {
    "ru": {
        "sheet_title": "Спецификация",
        "doc_title": "Спецификация конфигурации",
        "section_marker": "Русская версия",
        "meta_labels": {
            "configuration_id": "ID конфигурации",
            "project": "Проект",
            "contact": "Контакт",
            "contact_email": "Email контакта",
            "notes": "Примечания",
            "submitted_at": "Отправлено",
        },
        "kind_labels": {
            "equipment": "Оборудование",
            "module": "Модуль",
            "license": "Пакет лицензий",
            "service": "Сервис",
        },
        "table_header": [
            "Тип",
            "Название",
            "Количество",
            "ID родительского оборудования",
            "Целевые AP",
            "Единиц в пакете",
            "ID продукта",
            "ID модуля",
            "ID лицензии",
        ],
    },
    "en": {
        "sheet_title": "Specification",
        "doc_title": "Configuration specification",
        "section_marker": "English version",
        "meta_labels": {
            "configuration_id": "Configuration ID",
            "project": "Project",
            "contact": "Contact",
            "contact_email": "Contact email",
            "notes": "Notes",
            "submitted_at": "Submitted at",
        },
        "kind_labels": {
            "equipment": "Equipment",
            "module": "Module",
            "license": "License pack",
            "service": "Service",
        },
        "table_header": [
            "Type",
            "Name",
            "Quantity",
            "Parent equipment ID",
            "Target AP",
            "Units per pack",
            "Product ID",
            "Module ID",
            "License ID",
        ],
    },
}


def build_specification_from_configuration(db: Session, configuration_id: int) -> list[dict]:
    """Rebuild specification rows from persisted configuration items."""
    items = (
        db.query(ConfigurationItem)
        .filter(ConfigurationItem.configuration_id == configuration_id)
        .order_by(ConfigurationItem.id)
        .all()
    )
    rows: list[dict] = []
    for it in items:
        if it.product_id is not None and it.parent_product_id is None:
            p = db.query(Product).filter(Product.id == it.product_id).first()
            rows.append(
                {
                    "kind": "equipment",
                    "product_id": it.product_id,
                    "name": p.name if p else f"product#{it.product_id}",
                    "quantity": it.quantity,
                    "target_ap_count": None,
                }
            )
        elif it.module_id is not None:
            m = db.query(Module).filter(Module.id == it.module_id).first()
            rows.append(
                {
                    "kind": "module",
                    "module_id": it.module_id,
                    "name": m.name if m else f"module#{it.module_id}",
                    "parent_product_id": it.parent_product_id,
                    "quantity": it.quantity,
                }
            )
        elif it.license_id is not None:
            lic = db.query(License).filter(License.id == it.license_id).first()
            rows.append(
                {
                    "kind": "license",
                    "license_id": it.license_id,
                    "name": lic.name if lic else f"license#{it.license_id}",
                    "parent_product_id": it.parent_product_id,
                    "quantity": it.quantity,
                    "units_per_pack": lic.units_per_pack if lic else None,
                }
            )
        elif it.product_id is not None and it.parent_product_id is not None:
            p = db.query(Product).filter(Product.id == it.product_id).first()
            rows.append(
                {
                    "kind": "service",
                    "product_id": it.product_id,
                    "name": p.name if p else f"service#{it.product_id}",
                    "parent_product_id": it.parent_product_id,
                    "quantity": it.quantity,
                }
            )
    return rows


def _locale_text(locale: ExportLocale) -> dict[str, Any]:
    return _LOCALE_TEXT[locale]


def _project_meta(conf: Configuration, locale: ExportLocale) -> list[tuple[str, str]]:
    labels = _locale_text(locale)["meta_labels"]
    return [
        (labels["configuration_id"], str(conf.id)),
        (labels["project"], conf.project_name or ""),
        (labels["contact"], conf.project_contact_name or ""),
        (labels["contact_email"], conf.project_contact_email or ""),
        (labels["notes"], conf.project_notes or ""),
        (labels["submitted_at"], format_utc_as_moscow(conf.submitted_at)),
    ]


def _spec_table_header(locale: ExportLocale) -> list[str]:
    return list(_locale_text(locale)["table_header"])


def _kind_label(kind: str, locale: ExportLocale) -> str:
    labels = _locale_text(locale)["kind_labels"]
    return labels.get(kind, kind)


def _spec_row_cells(row: dict, locale: ExportLocale) -> list[Any]:
    kind = str(row.get("kind") or "")
    return [
        _kind_label(kind, locale),
        row.get("name") or "",
        row.get("quantity") if row.get("quantity") is not None else "",
        row.get("parent_product_id") if row.get("parent_product_id") is not None else "",
        row.get("target_ap_count") if row.get("target_ap_count") is not None else "",
        row.get("units_per_pack") if row.get("units_per_pack") is not None else "",
        row.get("product_id") if row.get("product_id") is not None else "",
        row.get("module_id") if row.get("module_id") is not None else "",
        row.get("license_id") if row.get("license_id") is not None else "",
    ]


def _write_spec_sheet(
    ws: Worksheet,
    *,
    conf: Configuration,
    specification: list[dict],
    locale: ExportLocale,
) -> None:
    text = _locale_text(locale)
    bold = Font(bold=True)
    header = _spec_table_header(locale)

    ws["A1"] = text["doc_title"]
    ws["A1"].font = bold
    row_idx = 3
    for label, value in _project_meta(conf, locale):
        ws.cell(row=row_idx, column=1, value=label).font = bold
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    row_idx += 1
    header_row = row_idx
    for col, title in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = bold
    row_idx += 1
    for spec_row in specification:
        for col, value in enumerate(_spec_row_cells(spec_row, locale), start=1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1

    for col in range(1, len(header) + 1):
        letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[letter].width = 18


def _specification_csv_text(
    *,
    conf: Configuration,
    specification: list[dict],
    locale: ExportLocale,
) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for label, value in _project_meta(conf, locale):
        writer.writerow([label, value])
    writer.writerow([])
    writer.writerow(_spec_table_header(locale))
    for spec_row in specification:
        writer.writerow(_spec_row_cells(spec_row, locale))
    return buf.getvalue()


def specification_to_xlsx_bytes(
    *,
    conf: Configuration,
    specification: list[dict],
) -> bytes:
    wb = Workbook()
    ru_ws = wb.active
    ru_ws.title = _LOCALE_TEXT["ru"]["sheet_title"][:31]
    _write_spec_sheet(ru_ws, conf=conf, specification=specification, locale="ru")

    en_ws = wb.create_sheet(title=_LOCALE_TEXT["en"]["sheet_title"][:31])
    _write_spec_sheet(en_ws, conf=conf, specification=specification, locale="en")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def specification_to_csv_bytes(
    *,
    conf: Configuration,
    specification: list[dict],
) -> bytes:
    """ZIP archive with separate Russian and English CSV files."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for locale in ("ru", "en"):
            csv_text = _specification_csv_text(
                conf=conf,
                specification=specification,
                locale=locale,
            )
            filename = f"configuration-{conf.id}-spec-{locale}.csv"
            zf.writestr(filename, csv_text.encode("utf-8-sig"))
    return buf.getvalue()
